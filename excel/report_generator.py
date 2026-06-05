"""
Excel report generator.
Produces a three-sheet workbook:
  1. Network Discovery
  2. Ethernet Channels
  3. Failed Devices
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from config import AppConfig
from network.discovery import DiscoveryResult
from network.cdp import CDPNeighbor
from network.etherchannel import EtherChannel

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates the Excel discovery report."""

    # ------------------------------------------------------------------
    # Style constants
    # ------------------------------------------------------------------
    _HDR_FONT = Font(name="Calibri", bold=True, color=AppConfig.COLOR_HEADER_FG, size=11)
    _HDR_FILL = PatternFill("solid", fgColor=AppConfig.COLOR_HEADER_BG)
    _HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

    _SUCCESS_FILL = PatternFill("solid", fgColor=AppConfig.COLOR_SUCCESS_BG)
    _FAILED_FILL  = PatternFill("solid", fgColor=AppConfig.COLOR_FAILED_BG)
    _ALT_FILL     = PatternFill("solid", fgColor=AppConfig.COLOR_ALT_ROW)
    _WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")

    _CELL_FONT    = Font(name="Calibri", size=10)
    _CELL_ALIGN   = Alignment(horizontal="left", vertical="center", wrap_text=False)
    _CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    _BORDER = Border(
        left  =Side(style="thin", color="D0D7DE"),
        right =Side(style="thin", color="D0D7DE"),
        top   =Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    # ------------------------------------------------------------------

    def __init__(self, results: List[DiscoveryResult], output_path: Path) -> None:
        self._results = results
        self._output_path = output_path

    def generate(self) -> None:
        """Build and save the workbook."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self._build_discovery_sheet(wb)
        self._build_etherchannel_sheet(wb)
        self._build_failed_sheet(wb)

        wb.save(self._output_path)
        logger.info("Report saved: %s", self._output_path)

    # ------------------------------------------------------------------
    # Sheet 1 — Network Discovery
    # ------------------------------------------------------------------

    def _build_discovery_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet(AppConfig.OUTPUT_SHEET_DISCOVERY)
        ws.row_dimensions[1].height = 22

        headers = [
            "IP", "Hostname", "VLANs", "SVI VLANs",
            "Neighbor", "Neighbor IP", "Platform",
            "Local Interface", "Remote Interface",
            "Gateway", "Status",
        ]
        self._write_header(ws, headers)

        row_idx = 2
        for result in self._results:
            if result.status == "SUCCESS" and result.cdp_neighbors:
                for neighbor in result.cdp_neighbors:
                    data = [
                        result.ip,
                        result.hostname,
                        ", ".join(result.vlans),
                        ", ".join(result.svi_interfaces),
                        neighbor.neighbor_hostname,
                        neighbor.neighbor_ip,
                        neighbor.platform,
                        neighbor.local_interface,
                        neighbor.remote_interface,
                        result.gateway,
                        result.status,
                    ]
                    fill = self._SUCCESS_FILL if (row_idx % 2 == 0) else self._ALT_FILL
                    self._write_row(ws, row_idx, data, fill)
                    row_idx += 1
            else:
                # No neighbors or failed — one row per device
                data = [
                    result.ip,
                    result.hostname or result.ip,
                    ", ".join(result.vlans),
                    ", ".join(result.svi_interfaces),
                    "", "", "", "", "",
                    result.gateway,
                    result.status,
                ]
                fill = self._FAILED_FILL if result.status == "FAILED" else self._SUCCESS_FILL
                self._write_row(ws, row_idx, data, fill)
                row_idx += 1

        self._apply_table(ws, headers, "TblDiscovery", "TableStyleMedium9",
                          row_idx - 1)
        self._freeze_and_filter(ws)
        self._auto_width(ws, headers)

        # Status column conditional color override
        status_col = len(headers)  # last column
        for r in range(2, row_idx):
            cell = ws.cell(row=r, column=status_col)
            if cell.value == "SUCCESS":
                cell.fill = PatternFill("solid", fgColor="00B050")
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            elif cell.value == "FAILED":
                cell.fill = PatternFill("solid", fgColor="C00000")
                cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    # ------------------------------------------------------------------
    # Sheet 2 — Ethernet Channels
    # ------------------------------------------------------------------

    def _build_etherchannel_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet(AppConfig.OUTPUT_SHEET_ETHERCHANNEL)
        ws.row_dimensions[1].height = 22

        headers = [
            "IP", "Hostname", "Port Channel", "Protocol",
            "Status", "Member Ports",
            "Neighbor", "Local Interface", "Remote Interface",
        ]
        self._write_header(ws, headers)

        row_idx = 2
        for result in self._results:
            if result.status != "SUCCESS" or not result.etherchannel:
                continue

            # Build a neighbor lookup: local_intf → neighbor
            neighbor_map: dict[str, CDPNeighbor] = {}
            for nbr in result.cdp_neighbors:
                if nbr.local_interface:
                    neighbor_map[nbr.local_interface] = nbr

            for ec in result.etherchannel:
                # Try to find a CDP neighbor linked to any member port
                linked_nbr: CDPNeighbor | None = None
                for port in ec.member_ports:
                    if port in neighbor_map:
                        linked_nbr = neighbor_map[port]
                        break

                data = [
                    result.ip,
                    result.hostname,
                    ec.port_channel,
                    ec.protocol,
                    ec.status,
                    ", ".join(ec.member_ports),
                    linked_nbr.neighbor_hostname if linked_nbr else "",
                    linked_nbr.local_interface if linked_nbr else "",
                    linked_nbr.remote_interface if linked_nbr else "",
                ]
                fill = self._SUCCESS_FILL if row_idx % 2 == 0 else self._ALT_FILL
                self._write_row(ws, row_idx, data, fill)
                row_idx += 1

        self._apply_table(ws, headers, "TblEtherChannel", "TableStyleMedium2",
                          row_idx - 1)
        self._freeze_and_filter(ws)
        self._auto_width(ws, headers)

    # ------------------------------------------------------------------
    # Sheet 3 — Failed Devices
    # ------------------------------------------------------------------

    def _build_failed_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet(AppConfig.OUTPUT_SHEET_FAILED)
        ws.row_dimensions[1].height = 22

        headers = ["IP", "Hostname", "Failure Reason", "Timestamp"]
        self._write_header(ws, headers)

        row_idx = 2
        for result in self._results:
            if result.status == "FAILED":
                data = [
                    result.ip,
                    result.hostname or "N/A",
                    result.failure_reason,
                    result.timestamp,
                ]
                self._write_row(ws, row_idx, data, self._FAILED_FILL)
                row_idx += 1

        self._apply_table(ws, headers, "TblFailed", "TableStyleMedium3",
                          row_idx - 1)
        self._freeze_and_filter(ws)
        self._auto_width(ws, headers)

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------

    def _write_header(self, ws: Worksheet, headers: List[str]) -> None:
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font    = self._HDR_FONT
            cell.fill    = self._HDR_FILL
            cell.alignment = self._HDR_ALIGN
            cell.border  = self._BORDER

    def _write_row(
        self, ws: Worksheet, row: int, data: List, fill: PatternFill
    ) -> None:
        ws.row_dimensions[row].height = 16
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = self._CELL_FONT
            cell.fill      = fill
            cell.alignment = self._CELL_ALIGN
            cell.border    = self._BORDER

    def _apply_table(
        self,
        ws: Worksheet,
        headers: List[str],
        table_name: str,
        table_style: str,
        last_row: int,
    ) -> None:
        if last_row < 2:
            return  # Need at least one data row for a valid table

        end_col = get_column_letter(len(headers))
        ref = f"A1:{end_col}{last_row}"

        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    def _freeze_and_filter(self, ws: Worksheet) -> None:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _auto_width(self, ws: Worksheet, headers: List[str]) -> None:
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header)

            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row,
                min_col=col_idx, max_col=col_idx,
            ):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
